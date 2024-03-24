"""
requirements:
msoffcrypto-tool
openpyxl
"""

import argparse
import datetime
import io
import json
import re
from collections import ChainMap
from itertools import islice
from typing import Any

import msoffcrypto
import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

version = "1.2.0"


def get_argument_parser(
    parent_parser: argparse.ArgumentParser | None = None,
) -> argparse.ArgumentParser:
    arg_parser = parent_parser or argparse.ArgumentParser(
        formatter_class=argparse.RawTextHelpFormatter,
        description="Tool used to extract data from an Excel spreadsheet to a json file.",
        epilog="",
    )
    arg_parser.add_argument("file_path", type=str, help="Excel spreadsheet path")
    arg_parser.add_argument(
        "-p",
        "--password",
        type=str,
        help="Password if the given xls/xlsx spreadsheet encrypted is encrypted.",
        required=False,
        default=None,
    )
    arg_parser.add_argument(
        "-s",
        "--sheet-name",
        metavar="sheet_name",
        type=str,
        help="""Name of the xls/xlsx sheet where the data are stored.""",
        required=False,
    )
    arg_parser.add_argument(
        "-o",
        "--output",
        type=str,
        help="""Output file path.""",
        required=False,
    )
    arg_parser.add_argument(
        "-P",
        "--pretty",
        type=bool,
        help="When active, the output json will be printed with a 2 character indentation.",
        required=False,
        default=False,
        action=argparse.BooleanOptionalAction,
    )
    format_group = arg_parser.add_mutually_exclusive_group(required=True)
    format_group.add_argument(
        "-f",
        "--format-file",
        metavar="format_file",
        type=str,
        help="""Used to provide a custom mapping format defined in a json file.
Example: {
  "row": "{_row_number}",
  "email": "{email}",
  "user": {
    "name": "{name}",
    "surname": "{surname}"
  },
  "operation_datetime": "{_now}"
}
The "{label}" notation searches for a field with that name in the header row.
The following fields are runtime generated:
- _row_number
- _now
Fields mapping supports the standard Python format notation for formatted strings.
""",
        required=False,
        default=None,
    )
    format_group.add_argument(
        "-m",
        "--mapping-labels",
        metavar="mapping_labels",
        type=str,
        help="""Used to provide a custom mapping labels list.""",
        nargs="+",
        required=False,
    )
    format_group.add_argument(
        "--flat",
        help="Maps a specific column as a json list of strings",
        type=str,
        required=False,
    )
    arg_parser.add_argument(
        "-u",
        "--unique-key",
        metavar="unique_key",
        help="""Used to provide the column number or label for the unique key.""",
        required=False,
    )
    return arg_parser


def decrypt(file_path: str, password: str) -> io.BytesIO:
    decrypted_workbook = io.BytesIO()

    with open(file_path, "rb") as file:
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted_workbook)

    return decrypted_workbook


def load_workbook(file_path: str, password: str | None = None) -> Workbook:
    return openpyxl.load_workbook(
        file_path if password is None else decrypt(file_path, password)
    )


def get_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name in workbook:
        worksheet = workbook[sheet_name]
        print(f"{type(worksheet) = }")
        return worksheet
    else:
        print(f"Worksheet not found: {sheet_name}")
        print(f"Available worksheet(s): {workbook.sheetnames}")
        exit(-1)


def find_column_index(row: tuple[Cell, ...], label: str) -> int:
    return next(i for i, cell in enumerate(row) if cell.value == label)


def load_format(format_file: str) -> dict:
    with open(format_file, "r") as file:
        return json.load(file)


def map_formatted_data(
    sheet: Worksheet,
    mapping_format: dict,
    unique_key: int | str | None,
) -> list[dict]:
    _now = datetime.datetime.now()

    def clean_label(label):
        return re.match(pattern=r"{(\w*)(:.*)?}", string=label).group(1)

    def extract_labels(mapping: dict[str, Any]) -> list[str]:
        labels = []
        for v in mapping.values():
            if type(v) is str:
                labels.append(clean_label(v))
            else:
                labels.extend(extract_labels(v))
        return labels

    label_indexes = {
        h.value: i
        for i, h in enumerate(next(sheet.rows))
        if h.value in extract_labels(mapping_format)
    }

    def _map_formatted(row: dict[str, Any], mf: dict | str) -> dict | str:
        return (
            mf.format(**row)
            if type(mf) is str
            else {k: _map_formatted(row, v) for k, v in mf.items()}
        )

    def map_formatted_row(row_number: int, row: tuple[Cell, ...], mf: dict) -> dict:
        return _map_formatted(
            dict(
                ChainMap(
                    {"_row_number": row_number, "_now": _now},
                    {label: row[index].value for label, index in label_indexes.items()},
                )
            ),
            mf=mf,
        )

    if unique_key is None:
        return [
            map_formatted_row(i, row, mapping_format)
            for i, row in islice(enumerate(sheet.rows), 1, None)
        ]
    else:
        unique_key_index = (
            unique_key if type(unique_key) is int else label_indexes[unique_key]
        )
        return list(
            {
                row[unique_key_index].value: map_formatted_row(i, row, mapping_format)
                for i, row in islice(enumerate(sheet.rows), 1, None)
            }.values()
        )


def map_flat(
    label: int | str,
    sheet: Worksheet,
) -> list[dict]:
    label_index = (
        label if type(label) is int else find_column_index(next(sheet.rows), label)
    )
    return list(set([row[label_index].value for row in islice(sheet.rows, 1, None)]))


def produce_output(
    file_path: str | None,
    data: list[dict[str, Any]],
    pretty: bool,
) -> None:
    if file_path is not None:
        with open(file_path, "w") as f:
            json.dump(data, fp=f, indent=2 if pretty else None)
    else:
        print(json.dumps(data, indent=2 if pretty else None))


def labels_to_format(mapping_labels: list[str]) -> dict:
    return {label: "{" + label + "}" for label in mapping_labels}


def _main(
    file_path: str,
    sheet_name: str | None,
    password: str | None,
    output: str | None,
    pretty: bool,
    mapping_labels: list[str] | None,
    unique_key: int | str | None,
    format_file: str | None,
    flat: str | None,
) -> None:
    workbook = load_workbook(file_path=file_path, password=password)
    sheet = get_sheet(workbook=workbook, sheet_name=sheet_name)

    if flat is not None:
        data = map_flat(
            sheet=sheet,
            label=flat,
        )
    else:
        data = map_formatted_data(
            sheet=sheet,
            mapping_format=(
                load_format(format_file=format_file)
                if format_file is not None
                else labels_to_format(mapping_labels=mapping_labels)
            ),
            unique_key=unique_key,
        )

    produce_output(file_path=output, data=data, pretty=pretty)


if __name__ == "__main__":
    argument_parser = get_argument_parser()
    argument_parser.add_argument(
        "-v", "--version", action="version", version=f"%(prog)s {version}"
    )
    args = argument_parser.parse_args()
    _main(**vars(args))
